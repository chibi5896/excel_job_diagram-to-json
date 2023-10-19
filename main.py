import openpyxl

def get_cells_within_border(workbook, sheet_name, border_style):
    # シートを選択
    sheet = workbook[sheet_name]

    # 罫線で囲まれたセルを格納するリスト
    cells_within_border = []

    for row in sheet.iter_rows():
        for cell in row:
            # セルの罫線スタイルを確認
            if (
                cell.border and
                (cell.border.top.style == border_style and
                cell.border.bottom.style == border_style and
                cell.border.left.style == border_style and
                cell.border.right.style == border_style)
            ):
                # セルが結合されている場合、結合セル内の左上のセルを取得
                cell_index = cell.coordinate
                for range_ in sheet.merged_cells.ranges:
                    merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
                    for row in merged_cells:
                        if cell_index in row:
                            cell = sheet[merged_cells[0][0]]
                            break
                cells_within_border.append(cell)

    return cells_within_border

def find_adjacent_cells(cells_within_border, cell):
    # セルの位置を取得
    position = cells_within_border.index(cell)
    prev_cell = None
    next_cell = None

    if position > 0:
        prev_cell = cells_within_border[position - 1]

    if position < len(cells_within_border) - 1:
        next_cell = cells_within_border[position + 1]

    return prev_cell, next_cell

# def get_merged_range(sheet, cell):
#     # セルが結合されているか確認
#     for merged_cell_range in sheet.merged_cells.ranges:
#         if cell.coordinate in merged_cell_range:
#             return merged_cell_range
#     return None

# def get_merged_value(sheet, merged_range):
#     # 結合セル内の値を取得
#     return sheet[merged_range.min_row][merged_range.min_col].value

def generate_mermaid_flowchart(cells_within_border, sheet):
    flowchart = "graph TD\n"

    for cell in cells_within_border:
        cell_value = cell.value

        prev_cell, next_cell = find_adjacent_cells(cells_within_border, cell)

        # if prev_cell:
        #     flowchart += f"{prev_cell.value} --> {cell_value}\n"

        if next_cell:
            flowchart += f"{cell_value} --> {next_cell.value}\n"

    return flowchart

def main():
    # エクセルファイルを読み込む
    workbook = openpyxl.load_workbook("data.xlsx")

    # シート名を指定
    sheet_name = "Sheet1"

    # 罫線のスタイルを指定
    border_style = "thin"

    # セルの結合を解除
    sheet = workbook[sheet_name]
    # for merged_range in sheet.merged_cells.ranges:
    #     sheet.unmerge_cells(merged_range.coord)

    # 罫線で囲まれたセルを取得
    cells_within_border = get_cells_within_border(workbook, sheet_name, border_style)

    # MermaidのFlowchartを生成
    flowchart = generate_mermaid_flowchart(cells_within_border, sheet)

    # Markdownファイルに出力
    with open("flowchart.md", "w", encoding="utf-8") as output_file:
        output_file.write("```mermaid\n")
        output_file.write(flowchart)
        output_file.write("```\n")

if __name__ == "__main__":
    main()
